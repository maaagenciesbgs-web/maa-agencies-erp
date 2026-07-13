from fastapi import APIRouter, UploadFile, File, HTTPException
from sqlmodel import select
from openpyxl import load_workbook

from database import get_session
from models import Product

router = APIRouter(
    prefix="/products",
    tags=["Products"],
)


@router.get("/")
def get_products():
    with get_session() as session:
        return session.exec(select(Product)).all()


@router.post("/")
def add_product(product: Product):
    with get_session() as session:
        session.add(product)
        session.commit()
        session.refresh(product)
        return {
            "message": "Product added successfully",
            "product": product,
        }


@router.put("/{product_id}")
def update_product(product_id: int, updated_product: Product):
    with get_session() as session:

        product = session.get(Product, product_id)

        if not product:
            raise HTTPException(
                status_code=404,
                detail="Product not found",
            )

        product.name = updated_product.name
        product.category = updated_product.category
        product.brand = updated_product.brand
        product.purchase_price = updated_product.purchase_price
        product.selling_price = updated_product.selling_price
        product.stock = updated_product.stock
        product.gst = updated_product.gst

        session.add(product)
        session.commit()
        session.refresh(product)

        return {
            "message": "Product updated successfully",
            "product": product,
        }


@router.delete("/{product_id}")
def delete_product(product_id: int):
    with get_session() as session:

        product = session.get(Product, product_id)

        if not product:
            raise HTTPException(
                status_code=404,
                detail="Product not found",
            )

        session.delete(product)
        session.commit()

        return {
            "message": "Product deleted successfully"
        }


@router.post("/import")
def import_products(file: UploadFile = File(...)):
    workbook = load_workbook(file.file)
    sheet = workbook.active

    imported = 0
    skipped = 0

    with get_session() as session:

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if not row:
                continue

            name = str(row[0] or "").strip()

            if name == "":
                continue

            # Skip totals / summary rows
            if (
                name.lower().startswith("total")
                or "scheme" in name.lower()
                or "cd amount" in name.lower()
            ):
                continue

            # Check duplicate by product name
            print("Checking:", name)
            existing = session.exec(
                select(Product).where(Product.name == name)
            ).first()

            print(existing)

            if existing:
                skipped += 1
                continue

            try:

                selling_price = float(row[1] or 0)
                purchase_price = float(row[2] or 0)
                stock = int(row[3] or 0)

            except:

                selling_price = 0
                purchase_price = 0
                stock = 0

            product = Product(
                name=name,
                category="General",
                brand="General",
                purchase_price=purchase_price,
                selling_price=selling_price,
                stock=stock,
                gst=18,
            )

            session.add(product)
            imported += 1

        session.commit()

    return {
        "status": "success",
        "imported": imported,
        "skipped": skipped,
        "message": f"Imported {imported} products. Skipped {skipped} duplicates."
    }